from random import random, randint

from django.conf import settings
from openpyxl import load_workbook

from .models import Client, Organization, Bill


def populate_client_table():
    client_sheet = load_workbook(settings.BASE_DIR / 'data/client_org.xlsx', read_only=True)['client']

    for row in list(client_sheet.values)[1:]:
        if not row or not row[0]:
            break
        Client.objects.create(name=row[0])


def populate_organization_table():
    organization_sheet = load_workbook(settings.BASE_DIR / 'data/client_org.xlsx', read_only=True)['organization']

    for row in list(organization_sheet.values)[1:]:
        if not row or not row[0]:
            break
        address = row[2]
        Organization.objects.create(
            name=row[1],
            client=Client.objects.get(name=row[0]),
            address=f'Адрес: {address}' if not address or address != '-' else address
        )


def populate_bills_table():
    bill_sheet = load_workbook(settings.BASE_DIR / 'data/bills.xlsx', read_only=True).active
    services = {1: 'консультация', 2: 'лечение', 3: 'стационар', 4: 'диагностика', 5: 'лаборатория'}

    for row in list(bill_sheet.values)[1:]:
        if not row or not row[0]:
            break
        random_service_class = randint(1, 5)
        Bill.objects.create(
            client=Client.objects.get(name=row[0]),
            organization=Organization.objects.get(name=row[1]),
            client_number=row[2],
            sum=row[3],
            date=row[4],
            fraud_score=random(),
            service_class=random_service_class,
            service_name=services[random_service_class]
        )

